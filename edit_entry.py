#!/usr/bin/python3.8

import sys
#import backend
import openpyxl as xl
from datetime import datetime, timedelta
import glob
from filelock import FileLock

SITEPATH = "/home/fvo/mysite/"
EVENTSPATH = SITEPATH+"events/"

def delete_row_content(starts,i):
    cols='BCDEFGHI'
    name= starts['B{}'.format(i)]
    for col in cols:
        starts['{}{}'.format(col,i)] = None
    print ("deleted row {}: {}".format(i,str(name)))

def save_change(file,col_letter,line,user_input):
    with FileLock(file+".lock"):
        wb = xl.load_workbook(file)
        starts = wb['StartList']
        starts['{}{}'.format(col_letter,line)] = user_input
        wb.save(file)
    print ("Change SAVED")

def swap(eventfile,time1,time2):
    '''swap the entrant for 2 specific times in the event file'''
    wb, event, fees, courses, starts = backend.load_sheets(event)
    entries = backend.get_entries(starts)
    entry1, entry2 = None,None
    for entry in entries:
        if entry["Start Time"]==time1: entry1=entry
        if entry["Start Time"]==time2: entry2=entry
    if entry1==None or entry2==None: return "Time not found"
    entry1["Start Time"]==time2
    entry2["Start Time"]==time1
    edit_entry(eventfile,**entry1)
    edit_entry(eventfile,**entry2)

if __name__ == "__main__":
    while True:
        full = False
        events = glob.glob(EVENTSPATH+'*.xlsx')
        e=1
        for event in events:
            print('[{}]...\t{}'.format(e,event))
            e +=1
        user_input = input('Select event number to edit..., add "f" for full e.g. 2f (otherwise blank lines ommitted), Q to quit ')
        user_input = str(user_input)
        if "Q" in user_input.upper(): break
        e = int(user_input[0])-1
        if "f" in user_input: full=True
        wb = xl.load_workbook(events[e])
        starts = wb['StartList']
        line = 2
        for row in starts.iter_rows(min_row=2, max_col=9, values_only=True):
            if full:
                if row[0] : print("[{}]...\t{:%H%M}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}".format(line,*row))
            else:
                if row[1] : print("[{}]...\t{:%H%M}\t{:.14}\t{}\t{:.10s}\t{}\t{}\t{}\t{:.10s}\t{:.8s}".format(line,*row))
            line +=1
        user_input = input('Select line number to edit..., or enter to quit ')
        if user_input == "": sys.exit(0)
        line = user_input
        cols = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        col=0
        for cell in starts[user_input]:
            print("[{}]...\t{}".format(cols[col],cell.value))
            col +=1
        user_input = input('Select column letter to edit..., or enter to quit ')
        col_letter=user_input.upper()
        if col_letter == "" or col_letter not in cols: sys.exit(0)
        user_input = input('Enter new value..., or enter to set to None ')
        if user_input == "": user_input=None
        starts['{}{}'.format(col_letter,line)] = user_input
        print("Confirm change to:")
        print("[{}]...\t{:%H%M}\t{}\t{}\t{}\t{}\n\t{}\t{}\t{}\t{}".format(line,*[c.value for c in starts[line]]))
        confirm_input = input('Enter Y or y to confirm ')
        if confirm_input.upper() == "Y":
            save_change(events[e],col_letter,line,user_input)
        else:
            print("Change abandonned")

"""
for file in glob.glob(EVENTSPATH+'*.xlsx'):
    i=2
    changed=False
    sixty_min_ago = datetime.now() - timedelta(hours=24)
    thirty_min_ago = datetime.now() - timedelta(hours=2)

"""


