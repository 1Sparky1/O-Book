# -*- coding: utf-8 -*-
"""
Created on Fri Jul  3 13:27:57 2020

@author: spark
"""

import openpyxl as xl
from datetime import datetime, timedelta
from filelock import FileLock
import logging
import math

CUTOFF_OFFSET = timedelta(days=-1, hours=+18) #6pm on day before event

def load_sheets(file):
    wb = xl.load_workbook(file)
    event = wb['EventDetails']
    fees = wb['EntryFees']
    courses = wb['Courses']
    starts = wb['StartList']
    #logging.info('all loaded - load_sheets in backend')
    return wb, event, fees, courses, starts

def get_members(file):
    wb = xl.load_workbook(file)
    members = wb['Members']
    members_dict = {}
    for row in members.iter_rows(min_row=2, max_col=2, values_only=True):
        members_dict[row[0]]=row[1]
    return members_dict

def strip_starts(start_times):
    ''' Returns a list of available start times '''
    ls = []
    for each in start_times.keys():
        if start_times[each] == None:
            ls.append(each)
    return ls

def read_ageclass(fees, event):
    ''' Reads the age classes from the Workbook and assigns their prices '''
    age_class = {}
    premium = 0
    if late_entries(event) : premium = get_event_details(event)["entry_premium"]
    for row in fees.iter_rows(min_row=1, max_col=2, values_only=True):
        #logging.info(row)
        if row[0]:
            if row[1] > 0:
                fee = row[1] + premium
            else:
                fee = row[1]
            age_class[row[0]] = fee
    return age_class

def read_course(event, courses):
    ''' Reads the courses from the Workbook and assigns their descriptions '''
    course = {}
    for row in courses.iter_rows(min_row=2, max_col=4, values_only=True):
        warning=""
        #logging.info(row)
        if row[1]:
            descrip = row[1]
            if row[3].upper()=="YES": warning = "<16s must be shadowed - "
        else:
            descrip = ""
        if not late_entries(event):
            course[row[0]] = warning+descrip
        elif not row[2] == 0:
            #only lists courses with maps available
            course[row[0]] = warning+descrip
    return course

def get_event_details(event):
    ''' Reads the event details from the Workbook and returns a dict '''
    details = {}

    details["hire_available"] = False
    details["members_only"] = False
    details["card_payments"] = False
    details['late_entries'] = False

    details["name"] = str(event['B1'].value)
    details["url"] = str(event['B2'].value)
    details["date"] = event['B3'].value
    details["date_pretty"] = (event['B3'].value).strftime('%A %-d %B %Y')
    details["organiser"] = str(event ['B4'].value)
    details["org_email"] = str(event ['B5'].value)
    details["closing"] = event['B6'].value
    details["closing_pretty"] = (event['B6'].value).strftime('%-d %b %Y %-I:%M%p')
    if str(event ['B7'].value).upper()=="YES": details["hire_available"]=True
    if str(event ['B8'].value).upper()=="YES": details["members_only"]=True
    if str(event ['B9'].value).upper()=="YES": details["card_payments"]=True
    if str(event ['B10'].value):
        details["late_entries"]=True
        details["entry_premium"]=event ['B10'].value
        # no idea why necessary - but force late entries to false when B10 is blank - fixed a bug!
        if event ['B10'].value == None: details["late_entries"]=False
    return details

def late_entries(event):
    details = get_event_details(event)
    cut_off = details["date"] + CUTOFF_OFFSET
    if all([details["late_entries"],details["closing"]<datetime.now(),datetime.now()<cut_off]):
        logging.info("Late ents set to true")
        return True
    else:
        return False


def event_closed(event):
    details = get_event_details(event)
    if details["closing"]<datetime.now():
        return True
    else:
        return False

def hire_available(event):
    details = get_event_details(event)
    return details["hire_available"]

def members_only(event):
    details = get_event_details(event)
    return details["members_only"]

def card_payments(event):
    details = get_event_details(event)
    return details["card_payments"]

def get_event_summary(event):
    ''' gets the event details then creates a HTML summary, and returns it'''
    details = get_event_details(event)
    return """  <strong>{name}</strong> on {date_pretty} <br>
                being organised by {organiser}<br>
                CLOSING DATE FOR ENTRIES:<br> {closing_pretty}<br>
                see <A Href={url}>{url}</A> for details<br>
            """.format(**details)

    course = {}
    for row in courses.iter_rows(min_row=1, max_col=2, values_only=True):
        course[row[0]] = row[1]
    return course

def read_time(starts):
    ''' Reads the start times from the Workbook and assigns anyone currently booked for that time '''
    time = {}
    for row in starts.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0]:
            temp = row[0].strftime("%H:%M")
            time[temp] = row[1]
    return time

def update_sheet(event, starts, courses, time, name, course, age_class, fee, dibber, phone, email, club):
    ''' Writes the information to the Start Sheet '''
    slot_available = False
    payment_warning=""
    for row in starts.iter_rows(min_row=2, max_col=8, values_only=True):
        if row[1]: #only consider rows with a name in the entry list, avoids an error as None.Upper() invalid
            if (row[1].upper() == name.upper() and row[3]== age_class) or (row[5] == dibber and dibber!="HIRE"):
                return False,'''This person appears to already have an entry, please <a href="javascript:history.back()">go back</a>
                        and enter a new person; make sure you are using the correct SI dibber number. You can check if you have already entered with the Entry List link at the top. If you are sure this is not a duplicate contact membership@fvo.org.uk.
                        If you have finished entering people <a href="/orienteering/invoice">go to your</a> invoice to finish.'''
    i=2
    for row in starts.iter_rows(min_row=2, max_col=9, values_only=True):
        if row[0] and row[0].strftime("%H:%M") == time and not row[1] :
            slot_available = True
            starts['B{}'.format(i)] = name
            starts['C{}'.format(i)] = course
            starts['D{}'.format(i)] = age_class
            starts['E{}'.format(i)] = fee
            starts['F{}'.format(i)] = dibber
            starts['G{}'.format(i)] = phone
            starts['H{}'.format(i)] = email
            starts['K{}'.format(i)] = club
            if fee>0:
                starts['I{}'.format(i)] = datetime.now()
            else:
                starts['I{}'.format(i)] = "FREE at {}".format(datetime.now())

            logging.info("status of lateentries {} and card_payments {}".format(late_entries(event),card_payments(event)))
            if late_entries(event) and "NO MAP" not in age_class.upper() :
                i=2
                for line in courses.iter_rows(min_row=2, max_col=4, values_only=True):
                    if line[0] == course:
                        current_value = courses['C{}'.format(i)].value
                        courses['C{}'.format(i)] = current_value -1
                        logging.info("Available maps on {} course reduced by 1".format(course))
                    i+=1

            if card_payments(event) : payment_warning = """<p><strong>This time will be reserved for the next 30 minutes -
                                  if you do not checkout on this device in the next 30 minutes,
                                  the time may be released.</strong><p>"""
            return slot_available,"""Time {} successfully reserved for {}. {}
                                    <p>Please <strong>note</strong> each participant - including shadowers must reserve a start time,
                                    multiple runners will not be accepted on a single entry. <p><small>Shadowers or family groups can still start
                                    at the same time on the day - but must reserve multiple slots for COVID compliance.</small></p>     """.format(time,name,payment_warning)
        i+=1

    #only reaches here is there are NO timeslots that match; this change enables more than one slot with same time.
    return slot_available,'Whilst you were completing the form this timeslot has been reserved by another participant, please <a href="javascript:history.back()">go back</a> and pick another.'

def confirmed_paid(file, entry_time, name, course):
    logging.info( 'Attempting to confirm payment in the xlsx' )
    wb2 = xl.load_workbook(file)
    starts = wb2['StartList']
    i=2
    recorded = False
    logging.info( 'before the for' )
    #logging.info(entry_time, name )
    for row in starts.iter_rows(min_row=2, max_col=9, values_only=True):
        excel_time = row[0].strftime("%H:%M")
        #logging.info(excel_time)
        if excel_time == entry_time and row[1] == name and row[2]==course:
            #logging.info( 'in the if' )
            starts['I{}'.format(i)] = "PAID at {}".format(datetime.now())
            logging.info( '{} paid for {} on {} at {}'.format( name,entry_time,file,datetime.now() ) )
            recorded = True
            break
        i+=1
    if not recorded:
        logging.info("###*** ENTRY NOT SAVED *** {} {} {} {} ***###".format(file, entry_time, name, course))
    write_wbook(wb2,file)
    logging.info("Saved")
    return recorded

def write_wbook(wb,filename):
    with FileLock(filename+".lock"):
        logging.info("Lock acquired.")
        wb.save(filename)
    logging.info("Lock released.")
    return

def get_entries(starts):
    entries=[]
    for row in starts.iter_rows(min_row=2, max_col=12, values_only=True):
        if row[1]:
            dct = {}
            dct['Name']=row[1]
            dct['Start Time']=row[0].strftime("%H:%M")
            dct['Course']=row[2]
            temp_age=row[3]
            dct['Age Class']=temp_age[0:3]
            dct['Dibber No.']=row[5]
            dct['Club']=row[10]
            entries.append(dct)
    return entries

def delete_row_content(starts,i):
    cols='BCDEFGHI'
    name= starts['B{}'.format(i)]
    for col in cols:
        starts['{}{}'.format(col,i)] = None
    logging.info("deleted row {}: {}".format(i,str(name)))

def cancel_entry(file, entry_time, name, course):
    i=2
    changed=False
    with FileLock(file+".lock"):
        wb = xl.load_workbook(file)
        starts = wb['StartList']
        for row in starts.iter_rows(min_row=2, max_col=9, values_only=True):
            if row[0]:
                excel_time = row[0].strftime("%H:%M")
            else:
                break
            if excel_time == entry_time and row[1] == name and row[2]==course:
                delete_row_content(starts,i)
                changed=True
            i+=1
        if changed:
            wb.save(file)
    logging.info("cancelled entry for {} on {}".format(name,course))

def get_age_class(age):
    age_class = 0
    if age <= 8:
        age_class = 10
    elif age <= 18:
        age_class = math.ceil(age/2) * 2
    elif age <= 29:
        age_class = 21
    elif age <= 79:
        age_class = math.floor(age/5) * 5
    else:
        age_class = 75
    if age_class > 0:
        return age_class
